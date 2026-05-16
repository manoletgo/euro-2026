"""Build the Euro 2026 itinerary as both .xlsx and .pdf from a single data source.

Run from the repo root:
    python scripts/build_itinerary.py

Outputs (overwritten on each run):
    output/euro_2026_itinerary.xlsx
    output/euro_2026_itinerary.pdf

When the source data in travel_info_data/info_notes.txt changes, update the
ROWS list below and rerun. Keep the schema, type categories, and colors in
sync with CLAUDE.md.
"""

from datetime import datetime
from pathlib import Path
from urllib.parse import quote
from xml.sax.saxutils import escape as xml_escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


HEADERS = [
    "Date",
    "Day",
    "Time (Depart)",
    "City / Route",
    "Type",
    "Location / Details",
    "Departure",
    "Arrival",
    "Duration",
    "Notes",
]

# Type prefix -> hex fill color. Used by both xlsx and pdf renderers.
TYPE_COLORS = {
    "Flight":          "DDEBF7",
    "Train":           "FFF2CC",
    "Lodging":         "E2EFDA",
    "Activity":        "FCE4D6",
    "Luggage storage": "EDEDED",
    "Tips":            "E4D7F2",
}

HEADER_FILL = "305496"

# Column widths (in mm for pdf, scaled for xlsx).
COL_WIDTHS_MM = [22, 12, 22, 32, 22, 60, 50, 50, 18, 70]


def gmaps_url(origin: str, destination: str, mode: str = "transit") -> str:
    """Build a Google Maps directions URL per the CLAUDE.md convention."""
    return (
        "https://www.google.com/maps/dir/?api=1"
        f"&origin={quote(origin)}"
        f"&destination={quote(destination)}"
        f"&travelmode={mode}"
    )


# Common origins reused across activity rows.
BOUNCE_ADDR = "Wiednergürtel 48/IV, 1040 Wien"
VIENNA_AIRBNB = "Margaretengürtel 6, 1050 Wien"
BOB_W_ADDR = "Bob W Stockholm Södermalm, Lundagatan, 117 27 Stockholm"
PRAGUE_AIRBNB = "Španělská 759/4, 120 00 Vinohrady, Praha"

ROWS = [
    ["2026-05-17", "Sun", "", "Manila → Istanbul", "Flight",
     "Flight J0085",
     "Ninoy Aquino International Airport (Manila) — 21:25",
     "Istanbul", "",
     "Departure from NAIA; international flight to Istanbul; in-flight overnight"],
    ["2026-05-18", "Mon", "", "Istanbul → Vienna", "Flight",
     "Flight J1883",
     "Istanbul", "Vienna — 09:15", "",
     "Arrival in Vienna"],
    ["2026-05-18", "Mon", "", "Vienna", "Luggage storage",
     "Bounce Luggage Storage – Vienna Prince Mobile (near Südtiroler Platz Station), Wiednergürtel 48/IV, Vienna 1040",
     "", "", "",
     "Drop luggage on arrival (~9:15 am); pick up later once Airbnb check-in is available"],
    ["2026-05-18", "Mon", "", "Vienna", "Tips",
     "Wiener Linien 72-hour pass (~€17.10) — U-Bahn + tram + bus across all zones",
     "", "", "",
     "Attractions: Sisi Ticket (€44) bundles Hofburg Imperial Apartments + Schönbrunn + Imperial Furniture; Vienna Pass for heavy sightseeing; Vienna State Opera standing-room (€4-15) "
     "| Explore: Sachertorte at Café Sacher, Wiener Schnitzel at Figlmüller, Naschmarkt food market, classic coffeehouses (Café Central, Hawelka)"],
    # Resolves <we're planning to go to st stephen's cahedral and followed by
    # Ausgrabungen michaelerplatz> placeholder. Split into two routed activity
    # rows per the Activity-routing convention.
    ["2026-05-18", "Mon", "10:00", "Vienna", "Activity",
     "St. Stephen's Cathedral (Stephansdom), Stephansplatz 3, 1010 Wien, Austria",
     f"From: {BOUNCE_ADDR} (Bounce Luggage)",
     "10:10 (est.)", "10 min",
     "Walk 3 min to Südtiroler Platz → U-Bahn U1 to Stephansplatz | "
     + gmaps_url(BOUNCE_ADDR, "Stephansplatz 3, 1010 Wien", "transit")],
    ["2026-05-18", "Mon", "11:30", "Vienna", "Activity",
     "Ausgrabungen Michaelerplatz, Michaelerplatz, 1010 Wien, Austria (open-air Roman & medieval excavations)",
     "From: Stephansplatz 3, 1010 Wien (St. Stephen's Cathedral)",
     "11:35 (est.)", "5 min",
     "Walk via Graben & Kohlmarkt | "
     + gmaps_url("Stephansplatz 3, 1010 Wien", "Michaelerplatz, 1010 Wien", "walking")],
    ["2026-05-18", "Mon", "16:30", "Vienna", "Lodging (Airbnb)",
     "Margaretengürtel 6, 1050 Wien, Austria",
     "Check-in: 2026-05-18 16:30", "Check-out: 2026-05-22", "4 nights",
     "Google Maps: https://maps.app.goo.gl/bAQYMCvSrB3vRvXn7"],
    # Resolves <2026-05-19 walking tour ... then Belvedere after lunch> placeholder.
    ["2026-05-19", "Tue", "09:45", "Vienna", "Activity",
     "GuruWalk 'Free Tour Vienna — Part 1, the highlights' (2-hour free/tip-based walking tour; stops include Heldenplatz, Hofburg, Stephansdom). Meeting point: Karlsplatz U-Bahn station, 1010 Wien, Austria",
     f"From: {VIENNA_AIRBNB} (Airbnb)",
     "10:10 (est.)", "25 min",
     "Tour 10:15–12:15. Booking: https://www.guruwalk.com/walks/36416-free-tour-vienna-part-1-the-highlights | U-Bahn U4 from Margaretengürtel → Karlsplatz | "
     + gmaps_url(VIENNA_AIRBNB, "Karlsplatz U-Bahn, 1010 Wien", "transit")],
    # Resolves <can you update belvedere palace row with details in the
    # may_19_vienna_belvedere_palace.jpg>. Source: that JPG screenshot —
    # Upper Belvedere Ticket, 19 May 2026 15:00, 2 x Adult, Belvedere Palace,
    # Prinz Eugen-Straße 27, 1030 Wien. Time (Depart) shifted to 14:30 to
    # arrive ~14:45 ahead of the 15:00 timed entry.
    ["2026-05-19", "Tue", "14:30", "Vienna", "Activity",
     "Upper Belvedere Palace — Prinz Eugen-Straße 27, 1030 Wien, Austria (Klimt's 'The Kiss'; Baroque palace + gardens)",
     "From: Karlsplatz / city centre (after walking tour 12:15 + lunch + free time)",
     "14:45 (est.)", "15 min",
     "Timed entry: 15:00, 2 adults (voucher in GetYourGuide / Klook app). Tram D from Karlsplatz → Belvedere (Schloss); arrive ~15 min before slot. After tour 12:15 and lunch, ~1h free before departure — gardens around Lower/Upper Belvedere are free | "
     + gmaps_url("Karlsplatz, 1010 Wien", "Upper Belvedere, Prinz Eugen-Straße 27, 1030 Wien", "transit")],
    ["2026-05-20", "Wed", "07:34", "Vienna → Budapest", "Train",
     "RegioJet RJ 1065",
     "Vienna Central Train Station — 07:34",
     "Budapest Déli — 10:14", "2h 40m",
     "Trainline: https://app.trainline.com/QrswMpOb12b"],
    ["2026-05-20", "Wed", "", "Budapest", "Tips",
     "BKV 24-hour ticket (~HUF 2,500 / €6.30) — metro + tram + bus; pair with M2 line from Déli station",
     "", "", "",
     "Attractions: Budapest Card (~€36/24h) bundles transit + 30+ sights + 2 thermal baths; Buda Castle district (Fisherman's Bastion outside is free); Széchenyi or Gellért Thermal Baths (bring swimsuit) "
     "| Explore: lángos at Karaván Street Food, goulash + kürtőskalács chimney cake, ruin bars in Jewish Quarter (Szimpla Kert), Parliament view from Pest embankment at sunset"],
    # Resolves <can you add this day-trip loop?> placeholder. The user's 12-step
    # loop is captured as 6 Activity rows (one per attraction stop); pure transit
    # and buffer steps are folded into the routing notes.
    ["2026-05-20", "Wed", "10:30", "Budapest", "Activity",
     "Hungarian Parliament Building (exterior) — Kossuth Lajos tér 1-3, 1055 Budapest",
     "From: Budapest Déli (Krisztina körút 37, 1013 Budapest)",
     "10:45 (est.)", "15 min",
     "M2 metro Déli → Kossuth tér (direct, ~10 min). 15-min photo stop on Kossuth tér; interior tours need advance booking | "
     + gmaps_url("Budapest Déli, Krisztina körút 37, 1013 Budapest", "Hungarian Parliament, Kossuth Lajos tér 1-3, 1055 Budapest", "transit")],
    ["2026-05-20", "Wed", "11:00", "Budapest", "Activity",
     "Shoes on the Danube Bank — Id. Antall József rkp., 1054 Budapest (memorial; free, outdoor)",
     "From: Hungarian Parliament, Kossuth Lajos tér",
     "11:05 (est.)", "5 min",
     "Walk south along Danube embankment | "
     + gmaps_url("Kossuth Lajos tér, Budapest", "Shoes on the Danube Bank, Budapest", "walking")],
    ["2026-05-20", "Wed", "11:20", "Budapest", "Activity",
     "St. Stephen's Basilica — Szent István tér 1, 1051 Budapest (climb the dome ~€4 for panorama)",
     "From: Shoes on the Danube Bank",
     "11:30 (est.)", "10 min",
     "Walk inland; ~45-min visit | "
     + gmaps_url("Shoes on the Danube Bank, Budapest", "St. Stephen's Basilica, Szent István tér 1, 1051 Budapest", "walking")],
    ["2026-05-20", "Wed", "12:15", "Budapest", "Activity",
     "Lunch — Vörösmarty / Erzsébet tér area, 1051 Budapest (suggested: Frici Papa or Belvárosi Disznótoros for goulash, chicken paprikash, kürtőskalács)",
     "From: St. Stephen's Basilica",
     "12:25 (est.)", "10 min",
     "Walk ~10 min; allow 1h for lunch | "
     + gmaps_url("St. Stephen's Basilica, Budapest", "Vörösmarty tér, Budapest", "walking")],
    ["2026-05-20", "Wed", "13:35", "Budapest", "Activity",
     "Buda Castle / Royal Palace courtyards — Szent György tér 2, 1014 Budapest (exterior + viewpoints over Pest)",
     "From: Vörösmarty / Erzsébet tér (post-lunch)",
     "13:55 (est.)", "20 min",
     "Walk Chain Bridge (~10 min, iconic Danube crossing) → Funicular up to Castle Hill (€6 one-way, ~5 min; or walk up ~15 min) | "
     + gmaps_url("Vörösmarty tér, Budapest", "Buda Castle, Szent György tér 2, 1014 Budapest", "walking")],
    ["2026-05-20", "Wed", "14:30", "Budapest", "Activity",
     "Fisherman's Bastion + Matthias Church — Szentháromság tér, 1014 Budapest (the iconic skyline shot of Pest)",
     "From: Buda Castle / Royal Palace courtyards",
     "14:35 (est.)", "5 min",
     "Short walk; ~45-min visit. After: Tram 19 or 41 from Clark Ádám tér south to Déli (~10 min); ~1.5h buffer at Déli before 17:45 RJ 1068 to Vienna | "
     + gmaps_url("Buda Castle, Budapest", "Fisherman's Bastion, Szentháromság tér, 1014 Budapest", "walking")],
    ["2026-05-20", "Wed", "17:45", "Budapest → Vienna", "Train",
     "RegioJet RJ 1068 (same-day return)",
     "Budapest Déli — 17:45",
     "Vienna Central Train Station — 20:27", "2h 42m",
     "Trainline: https://app.trainline.com/MeSYEbQb12b"],
    # Resolves <can you give options for attractions and expeditions?> placeholder.
    # Position infers: free day in Vienna, May 21 (between May 20 evening return
    # arrival 20:27 and May 22 06:39 departure to Prague). Belvedere dropped
    # from this list — now confirmed for May 19. Pick one (or two short ones)
    # and delete the rest.
    ["2026-05-21", "Thu", "09:30", "Vienna (Option 1/4)", "Activity",
     "Schönbrunn Palace & Gardens — Schönbrunner Schloßstraße 47, 1130 Wien (UNESCO Habsburg residence; Gloriette views)",
     f"From: {VIENNA_AIRBNB} (Airbnb)",
     "09:45 (est.)", "15 min",
     "U-Bahn U4 from Margaretengürtel → Schönbrunn (3 stops) | "
     + gmaps_url(VIENNA_AIRBNB, "Schönbrunn Palace, Vienna", "transit")],
    ["2026-05-21", "Thu", "10:00", "Vienna (Option 2/4)", "Activity",
     "Hofburg complex + Albertina Museum — Heldenplatz/Albertinaplatz 1, 1010 Wien (Imperial Apartments interior, Sisi Museum, Treasury; Albertina art museum 2 min walk). Distinct from May 19 walking-tour exterior pass.",
     f"From: {VIENNA_AIRBNB} (Airbnb)",
     "10:20 (est.)", "20 min",
     "U-Bahn U4 to Karlsplatz → walk 8 min | "
     + gmaps_url(VIENNA_AIRBNB, "Hofburg, Vienna", "transit")],
    ["2026-05-21", "Thu", "08:30", "Vienna → Bratislava (Option 3/4)", "Activity",
     "Day-trip to Bratislava, Slovakia — walkable Old Town + Bratislava Castle (~1h train each way; full day)",
     f"From: {VIENNA_AIRBNB} (Airbnb)",
     "09:45 (est.)", "1h 15m",
     "Transit to Wien Hauptbahnhof → REX 1 train to Bratislava hl. st. (~1h) | "
     + gmaps_url(VIENNA_AIRBNB, "Bratislava hl. st., Slovakia", "transit")],
    ["2026-05-21", "Thu", "08:00", "Vienna → Melk (Option 4/4)", "Activity",
     "Day-trip to Wachau Valley & Melk Abbey — Stift Melk, 3390 Melk (Danube wine region; Baroque abbey; ~1h train + optional Danube boat)",
     f"From: {VIENNA_AIRBNB} (Airbnb)",
     "09:15 (est.)", "1h 15m",
     "Transit to Wien Hauptbahnhof → ICE/RJ to Melk Bahnhof (~1h) | "
     + gmaps_url(VIENNA_AIRBNB, "Stift Melk, Melk", "transit")],
    ["2026-05-22", "Fri", "06:39", "Vienna → Prague", "Train",
     "RegioJet RJ 1030",
     "Vienna Central Train Station — 06:39",
     "Praha hl.n. — 10:56", "4h 17m",
     "Trainline: https://app.trainline.com/K0MINBRb12b"],
    # Resolves <can you put details here of the bounce booking ...> placeholder
    # from info_notes.txt. Source: forwarded Bounce confirmation .eml in
    # travel_info_data/. Single same-day booking: drop after train arrival,
    # pick up before heading to Vinohrady Airbnb.
    ["2026-05-22", "Fri", "", "Prague", "Luggage storage",
     "Bounce — Cafe Art of Alchemy (Prague Train Station), Opletalova 1418/23, 110 00 Nové Město, Prague",
     "Drop-off: 2026-05-22 11:00", "Pick-up: 2026-05-22 15:00", "4 hours",
     "2 regular bags, EUR 13.09 total. Booking ref: F13T68PD. Come INSIDE the café — luggage handled at the counter."],
    ["2026-05-22", "Fri", "", "Prague", "Tips",
     "DPP 72-hour pass (250 CZK / ~€10) — metro + tram + bus; tap-and-pay also accepted on trams",
     "", "", "",
     "Attractions: Prague Visitor Pass (or CoolPass) bundles Castle + Jewish Museum + tower climbs; Prague Castle basic-loop ticket (~CZK 250); book Old-New Synagogue + Klementinum tour online "
     "| Explore: trdelník (touristy but iconic), goulash + knedlíky at U Medvídků, Pilsner Urquell beer, Charles Bridge at dawn (avoid crowds), Vinohrady neighbourhood cafés near Airbnb"],
    ["2026-05-22", "Fri", "", "Prague", "Lodging (Airbnb)",
     "Španělská 759/4, 120 00 Vinohrady, Czechia",
     "Check-in: 2026-05-22", "Check-out: 2026-05-26", "4 nights",
     "Google Maps: https://maps.app.goo.gl/UcKWdy4NRJWvL3eZA"],
    ["2026-05-23", "Sat", "08:31", "Prague → Dresden", "Train",
     "Deutsche Bahn 178 (day-trip outbound)",
     "Praha hl.n. — 08:31", "Dresden Hbf — 10:50", "2h 19m",
     "Trainline: https://app.trainline.com/KXTU6ZSb12b"],
    ["2026-05-23", "Sat", "", "Dresden", "Tips",
     "DVB 1-day ticket (~€8.20) for trams + buses; Old Town (Altstadt) is compact and walkable from Hbf via tram 8/9",
     "", "", "",
     "Attractions: Dresden Museums Day Card (€25) covers Zwinger + Albertinum + Royal Palace; Frauenkirche entry free (donation); pre-book Historic Green Vault timed-entry online (sells out) "
     "| Explore: Eierschecke cake, Saxon white wine, Zwinger Old Masters Picture Gallery (Raphael's Sistine Madonna), Brühl's Terrace ('Balcony of Europe') for Elbe views"],
    # Resolves <can you add details of our river cruise tour? pls refer to the
    # [EXTERNAL]Fw_ Booking GYGMX4KYLANN confirmed _ Ticket instructions.eml file
    # in the travel_info_data folder>. Cruise: WEIßE FLOTTE SACHSEN GmbH, 1.5h
    # English-language sightseeing boat, departs 13:00 from piers 1-7 beneath
    # Brühlsche Terrasse (meeting point address: Augustusbrücke, Terrassenufer,
    # 01067 Dresden). Arrive 12:30 (boarding starts ~15 min before departure).
    ["2026-05-23", "Sat", "12:00", "Dresden", "Activity",
     "Dresden River Sightseeing Boat Cruise (WEIßE FLOTTE SACHSEN) — meeting point Augustusbrücke, Terrassenufer, 01067 Dresden (piers 1-7 beneath Brühlsche Terrasse)",
     "From: Dresden Hbf (after 10:50 train arrival)",
     "12:30 (est.)", "30 min",
     "Cruise 13:00–14:30 (1.5h, English audio commentary, 2 adults, €54). Booking ref: GYGMX4KYLANN; PIN: =U/7Fj9b. Operator: WEIßE FLOTTE SACHSEN GmbH (+49 351 866090). Tram 3 or 9 from Hbf → Theaterplatz/Synagoge, then short walk to Brühlsche Terrasse piers; arrive by 12:30 (boarding starts ~15 min before departure). Tickets in the GetYourGuide app | "
     + gmaps_url("Dresden Hbf", "Augustusbrücke, Terrassenufer, 01067 Dresden", "transit")],
    ["2026-05-23", "Sat", "19:10", "Dresden → Prague", "Train",
     "Deutsche Bahn 385 (day-trip return)",
     "Dresden Hbf — 19:10", "Praha hl.n. — 21:25", "2h 15m",
     "Trainline: https://app.trainline.com/Uo4dOhUb12b"],
    # Resolves <can you put in here details in may_24_prague_castle_guided_tour.jpg?
    # we'll be coming from the airbnb.>. Source: that JPG screenshot —
    # Prague Castle Guided Tour, Join-In Tour in English, 24 May 2026 11:00, 2 Adults,
    # booking ref HWG753830. Meeting point: Malostranská metro station exit,
    # Valdštejnská, 118 00 Praha 1 (small water fountain landmark out front).
    # Departure source per user note: the Vinohrady Airbnb.
    ["2026-05-24", "Sun", "10:30", "Prague", "Activity",
     "Prague Castle Guided Tour (Join-In, English) — meeting point Malostranská metro station exit, Valdštejnská, 118 00 Praha 1, Czechia",
     f"From: {PRAGUE_AIRBNB} (Airbnb)",
     "10:50 (est.)", "20 min",
     "Tour starts 11:00, 2 adults. Booking ref: HWG753830. Meet guide in front of the Malostranská metro exit (small water fountain landmark, covered in winter). Walk ~5 min to Náměstí Míru → Metro Line A (green) → Malostranská (4 stops, ~10 min) | "
     + gmaps_url(PRAGUE_AIRBNB, "Malostranská metro, Valdštejnská, 118 00 Praha 1", "transit")],
    ["2026-05-26", "Tue", "12:10", "Prague → Stockholm", "Flight",
     "Norwegian D82621 (Norwegian Air Sweden AOC AB); Ticket type: LOWFARE+; Seat 12C; Baggage: 1 checked + 1 underseat + 1 overhead",
     "Prague Václav Havel (PRG), Terminal 2 — 12:10",
     "Stockholm Arlanda (ARN), Terminal 5 — 14:05", "1h 55m",
     "Booking ref: Y9GHBZ; Ticket: 328 2404310661 (Norwegian Air Manolet.pdf)"],
    ["2026-05-26", "Tue", "", "Stockholm", "Tips",
     "SL Access 7-day pass (SEK 460 / ~€40) — metro + tram + bus + commuter rail + public ferries (incl. Route 80/82/89); airport: Arlanda Express train (~SEK 320 / €28, 18 min) or SL commuter rail (cheaper, ~40 min)",
     "", "", "",
     "Attractions: Vasa Museum (~SEK 220) is the single best stop; Skansen open-air museum + Royal Palace; Go City Stockholm Pass for heavy sightseeing "
     "| Explore: fika culture (coffee + kanelbulle cinnamon bun), Swedish meatballs at Pelikan or Tradition, Gamla Stan cobbled Old Town, Södermalm hipster cafés near hotel, archipelago boat from Strömkajen"],
    # Resolves <can you indicate in the Lodging (Hotel) Bob W Stockholm Södermalm
    # row departure and arrival? we depart from the airport, say 15 mins after
    # arrival of our flight>. Flight arrives ARN 14:05 → depart airport 14:20.
    # ARN → Bob W Södermalm via Arlanda Express (18 min) → metro to Mariatorget
    # (~12 min incl. walk + wait) → 5 min walk = arrive hotel ~15:15.
    ["2026-05-26", "Tue", "14:20", "Stockholm", "Lodging (Hotel)",
     "Bob W Stockholm Södermalm — Lundagatan, 117 27 Stockholm",
     "Stockholm Arlanda (ARN) — 14:20", "Bob W Södermalm — 15:15 (est.)", "4 nights",
     "Check-in: 2026-05-26 ~15:15; Check-out: 2026-05-30. Arlanda Express to Stockholm Central (~18 min, SEK 320) → metro T13/T14 to Mariatorget → 5 min walk. Google Maps: https://maps.app.goo.gl/QzqTafk4EjpTVc577"],
    # Resolves <after we check in at the hotel, we'll go ferry cruising. details
    # are in May 26 - Stockholm activity details.rtf...>. Source: that .rtf —
    # SL Route 82 from Slussen, ~1h round-trip Slussen → Skeppsholmen → Djurgården,
    # views of Gamla Stan, Royal Palace, Djurgården waterfront. Free with 7-day
    # SL pass (per the user's note in the .rtf). Then Tram 7 / metro to T-Centralen
    # for IKEA City inside Gallerian.
    ["2026-05-26", "Tue", "15:45", "Stockholm", "Activity",
     "SL public ferry Route 82 (sightseeing) — Slussen ferry terminal, Stadsgården 1, 116 45 Stockholm (round trip Slussen → Skeppsholmen → Djurgården; classic water views of Gamla Stan, Royal Palace, Djurgården waterfront, harbour islands)",
     f"From: {BOB_W_ADDR} (after check-in)",
     "16:00 (est.)", "15 min",
     "Walk ~10–15 min from hotel to Slussen ferry terminal (or 1 metro stop on T-bana to Slussen). Ferry covered by 7-day SL pass — €0 marginal cost. Round-trip ~50–60 min hits the user's ~1h sightseeing goal; alternatively hop off briefly at Djurgården before riding back | "
     + gmaps_url(BOB_W_ADDR, "Slussen ferry terminal, Stadsgården 1, 116 45 Stockholm", "walking")],
    ["2026-05-26", "Tue", "17:00", "Stockholm", "Activity",
     "IKEA City — Stockholm Gallerian, Hamngatan 37, 111 53 Stockholm (small-format city store inside Gallerian shopping mall, near T-Centralen)",
     "From: Slussen ferry terminal (after Route 82 round-trip)",
     "17:15 (est.)", "15 min",
     "Tram 7 from Djurgården (or metro from Slussen → T-Centralen, 2 stops, ~5 min) → 5 min walk into Gallerian. Covered by 7-day SL pass. Store typically closes 19:00 | "
     + gmaps_url("Slussen, Stockholm", "IKEA City, Gallerian, Hamngatan 37, 111 53 Stockholm", "transit")],
    ["2026-05-30", "Sat", "", "Stockholm → Istanbul", "Flight",
     "Flight J1796",
     "Stockholm Arlanda Airport", "Istanbul", "",
     "Transfer to Stockholm Arlanda; departure from Schengen area; in-flight overnight"],
    ["2026-05-31", "Sun", "", "Istanbul → Manila", "Flight",
     "Flight J0084",
     "Istanbul", "Manila", "",
     "Return flight to Manila"],
]

REPO_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_DIR = REPO_ROOT / "output"
PDF_PATH = OUTPUT_DIR / "euro_2026_itinerary.pdf"


def fill_for_type(type_val: str) -> str | None:
    for prefix, hex_color in TYPE_COLORS.items():
        if type_val.startswith(prefix):
            return hex_color
    return None


def build_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Itinerary"

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    body_font = Font(name="Arial", size=10)
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r_idx, row in enumerate(ROWS, start=2):
        type_val = row[4]
        hex_color = fill_for_type(type_val)
        fill = PatternFill("solid", fgColor=hex_color) if hex_color else None
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = body_font
            if fill:
                cell.fill = fill
            cell.alignment = wrap
            cell.border = border

    # Roughly map mm widths to Excel column units (~1 mm ≈ 0.55 unit).
    for i, mm_width in enumerate(COL_WIDTHS_MM, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(8, mm_width * 0.85)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_pdf(path: Path) -> None:
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        "cell", parent=styles["BodyText"],
        fontName="Helvetica", fontSize=8, leading=10, wordWrap="CJK",
    )
    header_style = ParagraphStyle(
        "header", parent=styles["BodyText"],
        fontName="Helvetica-Bold", fontSize=9, leading=11,
        textColor=colors.white, alignment=1,
    )
    title_style = ParagraphStyle(
        "title", parent=styles["Title"],
        fontName="Helvetica-Bold", fontSize=16, spaceAfter=8,
    )

    def cellp(text: str) -> Paragraph:
        if not text:
            return Paragraph("", cell_style)
        # Escape XML specials first (&, <, >) so URLs like
        # ?api=1&origin=... aren't mangled into &origin;=... by ReportLab's
        # entity parser. Insert <br/> after escaping so our line breaks survive.
        safe = xml_escape(str(text)).replace("\n", "<br/>")
        return Paragraph(safe, cell_style)

    table_data = [[Paragraph(h, header_style) for h in HEADERS]]
    table_data.extend([[cellp(c) for c in row] for row in ROWS])

    col_widths = [w * mm for w in COL_WIDTHS_MM]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{HEADER_FILL}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#BFBFBF")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])

    for i, row in enumerate(ROWS, start=1):
        hex_color = fill_for_type(row[4])
        if hex_color:
            style.add("BACKGROUND", (0, i), (-1, i), colors.HexColor(f"#{hex_color}"))

    table.setStyle(style)

    path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(path), pagesize=landscape(A3),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=12 * mm, bottomMargin=10 * mm,
        title="Euro 2026 Itinerary", author="Manolet",
    )
    doc.build([Paragraph("Euro 2026 Itinerary", title_style), Spacer(1, 4), table])


def main() -> None:
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    xlsx_path = OUTPUT_DIR / f"euro_2026_itinerary_{timestamp}.xlsx"

    build_xlsx(xlsx_path)
    build_pdf(PDF_PATH)

    print(f"Saved: {xlsx_path}  (new, retained as version history)")
    print(f"Saved: {PDF_PATH}  (overwritten)")
    print(f"Rows: {len(ROWS)}")


if __name__ == "__main__":
    main()
